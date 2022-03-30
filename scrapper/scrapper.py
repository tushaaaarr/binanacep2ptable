from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from time import sleep
from urllib.parse import urlparse, parse_qs
from openpyxl import Workbook
from elasticsearch import Elasticsearch
ELASTIC_SEARCH_HOST = "http://localhost:9200"
import os
driverpath = os.path.abspath(os.getcwd())+"/chromedriver.exe"
s=Service(executable_path= str(driverpath))

es = Elasticsearch(ELASTIC_SEARCH_HOST)
class App:
    def __init__(self):
        # selenium chrome driver configuration
        chrome_options = webdriver.ChromeOptions()
        chrome_options.add_argument("--start-maximized")
        chrome_options.add_argument("--headless")
        chrome_options.add_argument("--disable-web-security")
        chrome_options.add_argument("--allow-running-insecure-content")
        chrome_options.add_argument("--log-level=3")
        chrome_options.add_experimental_option("prefs", {"profile.managed_default_content_settings.images": 2,'profile.default_content_settings.images':2})
        self.driver = webdriver.Chrome(service = s,options=chrome_options)
        # self.data = []
        # self.wb = Workbook()
        # self.sheet = self.wb.active
        # self.es = Elasticsearch(ELASTIC_SEARCH_HOST)

    def scrap(self,link,index,pages)-> None:
        self.driver.get(link)
        sleep(3)
        try:
            self.driver.find_element(By.CLASS_NAME,"css-1pcqseb").click()
        except:
            pass
        page = 1
        count = 0
        for _ in range(pages):
            elements = self.driver.find_elements(By.CLASS_NAME,"css-1q1sp11")
            for element in elements:
                count += 1
                div = element.find_element(By.CLASS_NAME,"css-1rhb69f")
                advertiser = div.find_element(By.TAG_NAME,"a").text

                advertiser_href = div.find_element(By.TAG_NAME,"a").get_attribute('href')
                parsed_url = urlparse(advertiser_href)
                advertiser_no = parse_qs(parsed_url.query)['advertiserNo'][0]


                orders = element.find_element(By.CLASS_NAME,"css-1a0u4z7").text
                completion = element.find_element(By.CLASS_NAME,"css-19crpgd").text
                price = element.find_element(By.CLASS_NAME,"css-1m1f8hn").text
                fiat = element.find_element(By.CLASS_NAME,"css-dyl994").text
                div = element.find_element(By.CLASS_NAME,"css-vurnku").find_elements_by_class_name("css-vurnku")
                available = div[0].text
                limit = div[1].text.replace("\n","")

                payment_methods = []
                payment_cell = element.find_element(By.CLASS_NAME,"css-tlcbro")
                payment_boxes = payment_cell.find_elements(By.CLASS_NAME,"css-1n3cl9k")

                # print(payment_boxes)
                # return None
                for boxes in payment_boxes:
                    payment = boxes.find_elements(By.CSS_SELECTOR, "div")[1]
                    # print(payment.get_attribute('innerHTML'))
                    # return None
                    payment_methods.append(payment.get_attribute('innerHTML'))
                doc = {"exchange": "Binance","advertiser":advertiser,"orders":orders,"completion":completion,"price":price,"fiat":fiat,"payment":payment_methods,"available":available,"limit":limit}
                es.index(index=index,id=count,document=doc)
                # print(doc)
                # return None
                # self.sheet.cell(row=count,column=1).value = adveriser
                # self.sheet.cell(row=count,column=2).value = orders
                # self.sheet.cell(row=count,column=3).value = completion
                # self.sheet.cell(row=count,column=4).value = price
                # self.sheet.cell(row=count,column=5).value = fiat
                # self.sheet.cell(row=count,column=6).value = payment
                # self.sheet.cell(row=count,column=7).value = available
                # self.sheet.cell(row=count,column=8).value = limit
            page += 1
            buttons = self.driver.find_elements(By.TAG_NAME,"button")
            for b in buttons:
                if b.text==str(page):
                    b.click()
                    sleep(3)
                    break
    def close(self)-> None:
        # self.wb.save(filename="out.xlsx")
        self.driver.close()
a = App()
a.scrap("https://p2p.binance.com/en/trade/buy/?fiat=NGN&payment=ALL","advertisements",3)
a.close()

